import math
import io
from datetime import datetime
from flask import Blueprint, request, send_file
from app import db
from app.models.site import Site
from app.services.tank_stock_service import TankStockService
from app.utils import api_response
from app.models.tank import Tank
from app.models.fish_type import FishType

stock_bp = Blueprint("stock_bp", __name__, url_prefix="/api/tank-stock")

# # Apply login_required globally for all routes in this blueprint
# @stock_bp.before_request
# @login_required
# def before_request():
#     pass

# --------------------- Endpoints ---------------------
# List all tank stock records, with optional filtering by tankId or fishTypeId
@stock_bp.route("/paging", methods=["GET"])
def get_list_stock_paged():
    # Pagination
    page = request.args.get('page', 1, type=int)
    size = request.args.get('size', 10, type=int)

    # Sorting
    sort_field = request.args.get("sortField", "tank_name")
    sort_order = request.args.get("sortOrder", "asc")

    # Filters
    tank_id = request.args.get("tankId", type=int)
    fish_type_id = request.args.get("fishTypeId", type=int)
    site_id = request.args.get("siteId", type=int)
    search_text = request.args.get("searchText", "", type=str).strip()

    paginated_data = TankStockService.get_paged_tank_stock(
        page, size, sort_field, sort_order, tank_id, fish_type_id, site_id, search_text
    )

    return api_response("Tank stock retrieved successfully", data=paginated_data)

# List all tank stock records, with optional filtering by tankId or fishTypeId
@stock_bp.route("/", methods=["GET"])
def list_stock():
    tank_id      = request.args.get("tankId", type=int)
    fish_type_id = request.args.get("fishTypeId", type=int)
    site_id      = request.args.get("siteId", type=int)
    search_text  = request.args.get("searchText", "", type=str).strip()

    query = TankStockService.get_all_tank_stock(
        tank_id, fish_type_id, site_id, search_text
    )
    stock_list = query.all()

    # --- Python-side sort by tank_name then fish common_name ---
    stock_list.sort(
        key=lambda ts: (
            ts.tank.tank_name.lower()  if ts.tank and ts.tank.tank_name else "",
            ts.fish_type.common_name.lower() if ts.fish_type else "",
        )
    )

    return api_response(
        "Tank stock retrieved successfully",
        data=[ts.to_dict_light() for ts in stock_list],
    )

# Get a specific tank+fish_type stock record (read-only)
@stock_bp.route("/<int:tank_id>/<int:fish_type_id>", methods=["GET"])
def get_stock(tank_id, fish_type_id):
    stock = TankStockService.get_all_tank_stock(tank_id=tank_id, fish_type_id=fish_type_id)
    if not stock:
        return api_response("Tank stock not found", status=404)
    return api_response("Tank stock retrieved", data=stock.to_dict())

# List low-stock entries under a threshold (e.g., below 10)
@stock_bp.route("/low-stock", methods=["GET"])
def low_stock():
    threshold = request.args.get("threshold", default=10, type=int)
    status = request.args.get("status", default=None, type=str)
    site_id = request.args.get("siteId", type=int)
    
    # Fetch low stock entries using the service method with additional filters
    low_stocks = TankStockService.get_low_stock(threshold=threshold, status=status, site_id=site_id)
    
    # Convert to dict and return response
    result = [ts.to_dict() for ts in low_stocks]
    return api_response(f"Stocks with quantity below {threshold}", data=result)


# Get the total quantity of a specific fish type across all tanks
@stock_bp.route("/total-quantity/<int:fish_type_id>", methods=["GET"])
def get_total_quantity(fish_type_id):
    # Query to get the details of the fish type (e.g., common name, scientific name)
    fish_type = FishType.query.get(fish_type_id)
    if not fish_type:
        return api_response("Fish type not found", status=404)
    
    # Query the total quantity of the specific fish type across all tanks
    total_quantity = TankStockService.get_total_quantity_by_fish_type(fish_type_id)

    if total_quantity is None:
        return api_response("No records found for the given fish type", status=404)

    # Return the total quantity along with the fish type details
    return api_response(
        f"Total quantity of fish type {fish_type.common_name} : {total_quantity}",
        data={"totalQuantity": total_quantity, "fishType": fish_type.to_dict()}
    )

# Get total quantity for each tank, including a breakdown by fish type
@stock_bp.route("/summary", methods=["GET"])
def get_stock_summary_route():
    site_id = request.args.get("siteId", type=int)

    # Retrieve total quantity and fish breakdown via the service layer
    total_quantities = TankStockService.get_stock_summary(site_id)

    result = []
    for tank_id, total_quantity in total_quantities:
        tank = Tank.query.get(tank_id)
        if tank:
            result.append(TankStockService.build_tank_details(tank_id, total_quantity, tank))

    result.sort(key=lambda x: x['tankName'])

    return api_response(
        "Total quantities per tank with fish type breakdown retrieved successfully",
        data=result,
    )

@stock_bp.route("/dashboard/top-by-quantity", methods=["GET"])
def get_top_tanks_by_quantity():
    limit = request.args.get("limit", default=5, type=int)
    site_id = request.args.get("siteId", type=int)

    # Retrieve top tanks from service layer
    top_tanks = TankStockService.get_top_tanks_by_quantity(limit, site_id)

    result = []
    for tank_id, total_quantity in top_tanks:
        tank = Tank.query.get(tank_id)
        if tank:
            result.append(TankStockService.build_tank_details(tank_id, total_quantity, tank))

    return api_response("Top tanks by fish inventory", data=result)

# Get Site Distribution including Total Fish Count, Tank Count, and Fish Type Count
@stock_bp.route("/dashboard/site-distribution", methods=["GET"])
def get_site_distribution():
    site_id = request.args.get("siteId", type=int)

    # Retrieve site distribution from service layer
    site_distribution = TankStockService.get_site_distribution(site_id)

    result = []
    for site_id, tank_count, fish_type_count, total_fish_count in site_distribution:
        site = Site.query.get(site_id)
        if site:
            result.append({
                "siteId": site_id,
                "siteName": site.site_name,
                "tankCount": tank_count,
                "fishTypeCount": fish_type_count,
                "totalFishCount": total_fish_count or 0,
                "site": site.to_dict()
            })

    return api_response("Site Distribution Summary retrieved successfully", data=result)

@stock_bp.route("/fish-inventory", methods=["GET"])
def fish_inventory():
    site_id = request.args.get("siteId", type=int)
    tank_id = request.args.get("tankId", type=int)
    fish_type_id = request.args.get("fishTypeId", type=int)
    search_text = request.args.get("searchText", type=str)

    query = TankStockService.build_fish_inventory_query(site_id, tank_id, fish_type_id, search_text)
    query = query.order_by(FishType.common_name)

    fish_inventory = query.all()

    result = [
        {
            "typeId": fish[0],
            "commonName": fish[1],
            "typeCode": fish[2],
            "scientificName": fish[3],
            "totalTankCount": fish[4],
            "totalFishCount": fish[5]
        } for fish in fish_inventory
    ]

    return api_response("Fish Inventory retrieved successfully", data=result)

@stock_bp.route("/dashboard/top-by-fish-inventory", methods=["GET"])
def top_by_fish_inventory():
    site_id = request.args.get("siteId", type=int)
    tank_id = request.args.get("tankId", type=int)
    fish_type_id = request.args.get("fishTypeId", type=int)
    search_text = request.args.get("searchText", type=str)
    limit = request.args.get("limit", type=int, default=5)

    if not limit or limit < 5:
        limit = 5

    query = TankStockService.build_fish_inventory_query(site_id, tank_id, fish_type_id, search_text)
    query = query.order_by(db.desc("total_stock")).limit(limit)

    top_fish = query.all()

    result = [
        {
            "typeId": fish[0],
            "commonName": fish[1],
            "typeCode": fish[2],
            "scientificName": fish[3],
            "totalTankCount": fish[4],
            "totalFishCount": fish[5]
        } for fish in top_fish
    ]

    return api_response("Top Fish Inventory retrieved successfully", data=result)


@stock_bp.route("/fish-inventory/paging", methods=["GET"])
def fish_inventory_paged():
    page = request.args.get('page', 1, type=int)
    size = request.args.get('size', 10, type=int)
    site_id = request.args.get('siteId', type=int)
    tank_id = request.args.get("tankId", type=int)
    fish_type_id = request.args.get("fishTypeId", type=int)
    search_text = request.args.get("searchText", type=str)
    sort_field = request.args.get("sortField", "common_name")
    sort_order = request.args.get("sortOrder", "asc")

    valid_sort_fields = ["common_name", "type_code", "scientific_name"]
    if sort_field not in valid_sort_fields:
        sort_field = "common_name"

    query = TankStockService.build_fish_inventory_query(site_id, tank_id, fish_type_id, search_text)

    sort_mapping = {
        "common_name": FishType.common_name,
        "type_code": FishType.type_code,
        "scientific_name": FishType.scientific_name
    }
    sort_column = sort_mapping[sort_field]
    query = query.order_by(sort_column.desc() if sort_order == "desc" else sort_column)

    total_count = query.count()
    total_pages = math.ceil(total_count / size) if size > 0 else 0
    items = query.offset((page - 1) * size).limit(size).all()

    result = [
        {
            "typeId": fish[0],
            "commonName": fish[1],
            "typeCode": fish[2],
            "scientificName": fish[3],
            "totalTankCount": fish[4],
            "totalFishCount": fish[5]
        } for fish in items
    ]

    return api_response(
        "Fish Inventory retrieved successfully",
        data={
            "total": total_count,
            "totalPages": total_pages,
            "page": page,
            "size": size,
            "items": result
        }
    )

@stock_bp.route("/inventory/summary", methods=["GET"])
def inventory_summary():
    site_id = request.args.get("siteId", type=int)
    summary_data = TankStockService.get_fish_summary(site_id, detailed=False)
    return api_response("Inventory Summary", data=summary_data)

@stock_bp.route("/dashboard/summary", methods=["GET"])
def dashboard_summary():
    site_id = request.args.get("siteId", type=int)
    summary_data = TankStockService.get_fish_summary(site_id, detailed=True)
    return api_response("Dashboard Summary", data=summary_data)

@stock_bp.route("/dashboard/total-fish-count", methods=["GET"])
def dashboard_total_fish_count():
    site_id = request.args.get("siteId", type=int)
    summary_data = TankStockService.get_fish_summary(site_id)
    return api_response("Total Fish Count", data=summary_data)

@stock_bp.route("/dashboard/total-active-tanks", methods=["GET"])
def active_tanks():
    site_id = request.args.get("siteId", type=int)

    # Use the dedicated method for active tank count
    active_tank_count = TankStockService.get_active_tank_count(site_id)

    return api_response("Total Active Tanks", data={
        "totalTankCount": active_tank_count,
        "label": "Available for fish storage"
    })

@stock_bp.route("/dashboard/tanks-with-stock", methods=["GET"])
def tanks_with_stock():
    site_id = request.args.get("siteId", type=int)

    # Use the service layer to get the tank stock data
    tank_stock_data = TankStockService.get_tanks_with_stock(site_id)

    return api_response("Tanks With Stock", data=tank_stock_data)


@stock_bp.route("/overall-stock-status/export", methods=["GET"])
def export_overall_stock_status():
    site_id = request.args.get("siteId", type=int)
    records = TankStockService.get_overall_stock_status(site_id)

    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "OverallStockStatus"

    # Header section with metadata
    title_cell = ws["A1"]
    title_cell.value = "Overall Stock Status"
    title_cell.font = Font(bold=True)
    ws.merge_cells("A1:G1")
    ws["A2"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"

    site_label = "All Sites"
    if site_id:
        site_name = db.session.query(Site.site_name).filter_by(site_id=site_id).scalar()
        site_label = site_name or f"Site {site_id}"
    ws["A3"] = f"Site: {site_label}"

    ws.append([])  # blank row before table headers
    ws.append([
        "SiteName",
        "TankCode",
        "TankName",
        "FishTypeCode",
        "FishTypeName",
        "Quantity",
        "LastUpdated",
    ])

    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    prev_site = None
    prev_tank = None
    for (
        site_name,
        tank_code,
        tank_name,
        type_code,
        common_name,
        quantity,
        last_updated,
    ) in records:
        show_group = not (site_name == prev_site and tank_code == prev_tank)
        ws.append([
            site_name if show_group else "",
            tank_code if show_group else "",
            tank_name if show_group else "",
            type_code,
            common_name,
            quantity,
            last_updated.strftime("%Y-%m-%d %H:%M:%S") if last_updated else None,
        ])
        prev_site, prev_tank = site_name, tank_code

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return send_file(
        stream,
        as_attachment=True,
        download_name="overall_stock_status.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
